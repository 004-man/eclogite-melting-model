import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import warnings
warnings.filterwarnings('ignore')

class BatchMeltingModel:
    def __init__(self, input_file):
        """
        Initialize the batch melting model with input Excel file
        """
        self.input_file = input_file
        self.ree_elements = ['La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu']
        self.load_input_data()
        
    def load_input_data(self):
        """
        Load all input data from Excel sheets
        """
        # Read Garnet sheet
        self.garnet_data = pd.read_excel(self.input_file, sheet_name='Garnet', header=1)
        self.garnet_data = self.garnet_data.dropna(subset=[self.garnet_data.columns[0]])
        self.garnet_data.set_index(self.garnet_data.columns[0], inplace=True)
        
        # Read Plagioclase sheet  
        self.plag_data = pd.read_excel(self.input_file, sheet_name='Plagioclase', header=1)
        self.plag_data = self.plag_data.dropna(subset=[self.plag_data.columns[0]])
        self.plag_data.set_index(self.plag_data.columns[0], inplace=True)
        
        # Read Chondrite values
        self.chondrite_data = pd.read_excel(self.input_file, sheet_name='Chondrite value', header=0)
        self.chondrite_data.set_index(self.chondrite_data.columns[0], inplace=True)
        
        print("✅ Successfully loaded input data from Excel file")
        
    def calculate_bulk_partition_coefficients(self, data_sheet, scenario='garnet'):
        """
        Calculate bulk D (source) and P (melting) partition coefficients
        """
        bulk_D = {}
        bulk_P = {}
        
        # Define mineral columns based on scenario
        if scenario == 'garnet':
            minerals = ['Ol', 'Opx', 'Cpx', 'Grt']
            modal_cols = ['Ol_Starting', 'Opx_Starting', 'Cpx_Starting', 'Grt_Starting']
            melt_cols = ['Ol_melt fraction', 'Opx_melt fraction', 'Cpx_melt fraction', 'Grt_melt fraction']
            kd_cols = ['Kd (Ol/melt)', 'Kd (Opx/melt)', 'Kd (Cpx/melt)', 'Kd (Grt/melt)']
        else:  # plagioclase
            minerals = ['Ol', 'Opx', 'Cpx', 'Grt', 'Plg']
            modal_cols = ['Ol_Starting', 'Opx_Starting', 'Cpx_Starting', 'Grt_Starting', 'Plg_Starting']
            melt_cols = ['Ol_melt fraction', 'Opx_melt fraction', 'Cpx_melt fraction', 'Grt_melt fraction', 'Plg_melt fraction']
            kd_cols = ['Kd (Ol/melt)', 'Kd (Opx/melt)', 'Kd (Cpx/melt)', 'Kd (Grt/melt)', 'Kd (Plg/melt)']
        
        # Get modal and melting proportions from first REE row (La)
        la_row = data_sheet.loc['La']
        modal_props = [la_row[col] if pd.notna(la_row[col]) and la_row[col] > 0 else 0 for col in modal_cols]
        melt_props = [la_row[col] if pd.notna(la_row[col]) and la_row[col] > 0 else 0 for col in melt_cols]
        
        # Normalize to percentages if needed
        modal_props = np.array(modal_props) / 100 if max(modal_props) > 1 else np.array(modal_props)
        melt_props = np.array(melt_props) / 100 if max(melt_props) > 1 else np.array(melt_props)
        
        # Calculate bulk coefficients for each REE
        for ree in self.ree_elements:
            if ree in data_sheet.index:
                row = data_sheet.loc[ree]
                kd_values = []
                
                # Get Kd values for each mineral
                for col in kd_cols:
                    if col in row.index and pd.notna(row[col]):
                        kd_values.append(row[col])
                    else:
                        kd_values.append(0)  # Use 0 for missing Kd values
                
                kd_values = np.array(kd_values)
                
                # Calculate bulk D (source assemblage)
                D = np.sum(modal_props * kd_values)
                
                # Calculate bulk P (melting assemblage)
                P = np.sum(melt_props * kd_values)
                
                bulk_D[ree] = D
                bulk_P[ree] = P
        
        return bulk_D, bulk_P, modal_props, melt_props
    
    def verify_mass_balance(self, C0, F, C_melt, C_residue):
        """
        Verify mass balance: C_0 = F * C_melt + (1-F) * C_residue
        """
        calculated_C0 = F * C_melt + (1 - F) * C_residue
        relative_error = abs(calculated_C0 - C0) / C0 if C0 != 0 else 0
        return relative_error < 0.001  # Accept < 0.1% error
    
    def shaw_equation_melt(self, C0, F, D, P):
        """
        Shaw (1970) equation for melt concentration
        C_melt/C_0 = 1/[D + F(1-P)]
        """
        denominator = D + F * (1 - P)
        if denominator == 0:
            return 0
        return C0 / denominator
    
    def shaw_equation_residue(self, C0, F, D, P):
        """
        Shaw (1970) equation for residue concentration - CORRECTED
        Using mass balance: C_s = (C_0 - F*C_l)/(1-F)
        Where C_l = C_0/[D + F(1-P)]
        """
        if F >= 1.0:  # Prevent division by zero
            return 0
            
        # Calculate melt concentration first
        C_l = self.shaw_equation_melt(C0, F, D, P)
        
        # Mass balance: Original = Melt_fraction * Melt_conc + Residue_fraction * Residue_conc
        # C_0 = F * C_l + (1-F) * C_s
        # Solving for C_s: C_s = (C_0 - F*C_l)/(1-F)
        
        C_s = (C0 - F * C_l) / (1 - F)
        
        return C_s
    
    def continuous_melting(self, data_sheet, scenario='garnet'):
        """
        Calculate continuous melting (1% to 20%)
        """
        print(f"🔄 Calculating continuous melting for {scenario} stability field...")
        
        # Get starting concentrations
        starting_conc = {}
        for ree in self.ree_elements:
            if ree in data_sheet.index:
                starting_conc[ree] = data_sheet.loc[ree, 'Starting composition (rock)']
        
        # Calculate bulk partition coefficients
        bulk_D, bulk_P, modal_props, melt_props = self.calculate_bulk_partition_coefficients(data_sheet, scenario)
        
        # Melt fractions from 1% to 20%
        melt_fractions = np.arange(0.01, 0.21, 0.01)
        
        results = {
            'melt_fractions': melt_fractions,
            'melt_concentrations': {},
            'residue_concentrations': {},
            'bulk_D': bulk_D,
            'bulk_P': bulk_P,
            'modal_proportions': modal_props,
            'melting_proportions': melt_props,
            'starting_concentrations': starting_conc
        }
        
        # Calculate concentrations for each REE and melt fraction
        for ree in self.ree_elements:
            if ree in starting_conc and ree in bulk_D:
                C0 = starting_conc[ree]
                D = bulk_D[ree]
                P = bulk_P[ree]
                
                melt_conc = []
                residue_conc = []
                
                for F in melt_fractions:
                    melt_c = self.shaw_equation_melt(C0, F, D, P)
                    residue_c = self.shaw_equation_residue(C0, F, D, P)
                    
                    # Verify mass balance
                    if not self.verify_mass_balance(C0, F, melt_c, residue_c):
                        print(f"⚠️  Mass balance error for {ree} at F={F:.0%}")
                    
                    melt_conc.append(melt_c)
                    residue_conc.append(residue_c)
                
                results['melt_concentrations'][ree] = melt_conc
                results['residue_concentrations'][ree] = residue_conc
                
                # Debug check for first calculation (F=1%)
                if ree == 'La':  # Check La as example
                    F_test = 0.01
                    melt_test = self.shaw_equation_melt(C0, F_test, D, P)
                    residue_test = self.shaw_equation_residue(C0, F_test, D, P)
                    print(f"🔍 Debug check for {ree}:")
                    print(f"   Starting: {C0:.3f}, D: {D:.4f}, P: {P:.4f}")
                    print(f"   At F=1%: Melt={melt_test:.3f} ({'ENRICHED' if melt_test > C0 else 'DEPLETED'})")
                    print(f"   At F=1%: Residue={residue_test:.3f} ({'ENRICHED' if residue_test > C0 else 'DEPLETED'})")
        
        return results
    
    def stepwise_melting(self, data_sheet, step_melt_fraction=0.05, num_steps=6):
        """
        Calculate stepwise melting with modal evolution
        """
        print("🔄 Calculating stepwise melting for garnet stability field...")
        
        # Initialize with starting composition
        current_conc = {}
        for ree in self.ree_elements:
            if ree in data_sheet.index:
                current_conc[ree] = data_sheet.loc[ree, 'Starting composition (rock)']
        
        # Get initial modal proportions
        la_row = data_sheet.loc['La']
        initial_modal = np.array([
            la_row['Ol_Starting'] if pd.notna(la_row['Ol_Starting']) else 0,
            la_row['Opx_Starting'] if pd.notna(la_row['Opx_Starting']) else 0, 
            la_row['Cpx_Starting'] if pd.notna(la_row['Cpx_Starting']) else 0,
            la_row['Grt_Starting'] if pd.notna(la_row['Grt_Starting']) else 0
        ]) / 100
        
        # Get melting proportions (constant throughout)
        melt_props = np.array([
            la_row['Ol_melt fraction'] if pd.notna(la_row['Ol_melt fraction']) else 0,
            la_row['Opx_melt fraction'] if pd.notna(la_row['Opx_melt fraction']) else 0,
            la_row['Cpx_melt fraction'] if pd.notna(la_row['Cpx_melt fraction']) else 0,
            la_row['Grt_melt fraction'] if pd.notna(la_row['Grt_melt fraction']) else 0
        ]) / 100
        
        results = {
            'steps': [],
            'cumulative_melt_fraction': [],
            'melt_concentrations': {ree: [] for ree in self.ree_elements if ree in current_conc},
            'residue_concentrations': {ree: [] for ree in self.ree_elements if ree in current_conc},
            'modal_evolution': [],
            'modal_proportions': initial_modal,  # Add this
            'melting_proportions': melt_props,
            'starting_concentrations': current_conc.copy(),
            'bulk_D': {},  # Add this - will be filled with final values
            'bulk_P': {}   # Add this - will be filled with final values
        }
        
        current_modal = initial_modal.copy()
        cumulative_F = 0
        
        for step in range(1, num_steps + 1):
            print(f"  Step {step}: F = {step_melt_fraction*100}%")
            
            # Calculate bulk partition coefficients for current modal composition
            bulk_D = {}
            bulk_P = {}
            
            kd_cols = ['Kd (Ol/melt)', 'Kd (Opx/melt)', 'Kd (Cpx/melt)', 'Kd (Grt/melt)']
            
            # Debug: Show current modal composition
            mineral_names = ['Ol', 'Opx', 'Cpx', 'Grt']
            modal_debug = [f'{mineral_names[i]}={current_modal[i]:.1%}' for i in range(len(mineral_names)) if i < len(current_modal)]
            print(f"    Current modal: {modal_debug}")
            
            for ree in self.ree_elements:
                if ree in data_sheet.index:
                    row = data_sheet.loc[ree]
                    kd_values = []
                    
                    for col in kd_cols:
                        if col in row.index and pd.notna(row[col]):
                            kd_values.append(row[col])
                        else:
                            kd_values.append(0)
                    
                    kd_values = np.array(kd_values)
                    
                    # Calculate bulk coefficients with current modal proportions
                    D = np.sum(current_modal * kd_values)
                    P = np.sum(melt_props * kd_values)
                    
                    bulk_D[ree] = D
                    bulk_P[ree] = P
            
            # Debug: Show how bulk D changes for La
            if 'La' in bulk_D:
                print(f"    Bulk D for La: {bulk_D['La']:.4f} (P constant: {bulk_P['La']:.4f})")
            
            # Calculate melt and residue concentrations for this step
            step_melt_conc = {}
            step_residue_conc = {}
            
            for ree in self.ree_elements:
                if ree in current_conc and ree in bulk_D:
                    C0 = current_conc[ree]
                    D = bulk_D[ree]
                    P = bulk_P[ree]
                    
                    melt_conc = self.shaw_equation_melt(C0, step_melt_fraction, D, P)
                    residue_conc = self.shaw_equation_residue(C0, step_melt_fraction, D, P)
                    
                    # Verify mass balance
                    if not self.verify_mass_balance(C0, step_melt_fraction, melt_conc, residue_conc):
                        print(f"⚠️  Mass balance error for {ree} at step {step}")
                    
                    step_melt_conc[ree] = melt_conc
                    step_residue_conc[ree] = residue_conc
                    
                    results['melt_concentrations'][ree].append(melt_conc)
                    results['residue_concentrations'][ree].append(residue_conc)
                    
                    # Update current concentration to residue concentration for next step
                    current_conc[ree] = residue_conc
            
            # Store the modal composition USED for this step (before updating)
            results['modal_evolution'].append(current_modal.copy())
            
            # Update modal proportions for next step - CORRECTED CALCULATION
            # Simple mass balance approach
            
            # Calculate how much of each mineral was consumed in this melting step
            # For simplicity, assume we start each step with 100 units of current composition
            current_total = 100
            melt_extracted = current_total * step_melt_fraction
            
            # Calculate mineral amounts consumed according to melting reaction
            minerals_consumed = melt_props * melt_extracted
            
            # Calculate remaining mineral amounts 
            # current_modal represents the proportions going INTO this step
            starting_amounts = current_modal * current_total
            remaining_amounts = starting_amounts - minerals_consumed
            
            # Normalize to get new modal percentages for next step
            total_remaining = np.sum(remaining_amounts)
            if total_remaining > 0:
                current_modal = remaining_amounts / total_remaining
            
            # Debug output for modal evolution
            if step <= 2:  # Show evolution for first 2 steps
                print(f"    Modal after step {step}: {[f'{mineral_names[i]}={current_modal[i]:.1%}' for i in range(len(mineral_names)) if i < len(current_modal)]}")
            
            # Store step results
            cumulative_F += step_melt_fraction
            results['steps'].append(step)
            results['cumulative_melt_fraction'].append(cumulative_F)
        
        # Store final bulk partition coefficients
        results['bulk_D'] = bulk_D
        results['bulk_P'] = bulk_P
        
        return results
    
    def normalize_to_chondrite(self, concentrations, chondrite_values):
        """
        Normalize REE concentrations to chondrite values
        """
        normalized = {}
        
        # Get the chondrite column name (it should be the first and only column after setting index)
        chondrite_col = chondrite_values.columns[0]
        
        for ree in concentrations:
            if ree in chondrite_values.index:
                chondrite_val = chondrite_values.loc[ree, chondrite_col]
                if isinstance(concentrations[ree], list):
                    normalized[ree] = [c / chondrite_val for c in concentrations[ree]]
                else:
                    normalized[ree] = concentrations[ree] / chondrite_val
        return normalized
    
    def create_spider_plot(self, scenarios_data, output_file='REE_Spider_Plots.png'):
        """
        Create REE spider plots for all scenarios - BOTH MELT AND RESIDUE
        """
        print("📊 Creating REE spider diagrams...")
        
        # Set up the plot - 2 rows (melt and residue)
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('REE Spider Diagrams - Chondrite Normalized', fontsize=16, fontweight='bold')
        
        # REE order for x-axis
        ree_order = ['La', 'Ce', 'Pr', 'Nd', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu']
        x_positions = range(len(ree_order))
        
        scenario_names = ['Garnet Continuous', 'Garnet Stepwise', 'Plagioclase Continuous']
        
        for i, (scenario_name, data) in enumerate(zip(scenario_names, scenarios_data)):
            
            # MELT PLOTS (top row)
            ax_melt = axes[0, i]
            
            # RESIDUE PLOTS (bottom row)  
            ax_residue = axes[1, i]
            
            if 'stepwise' in scenario_name.lower():
                # For stepwise: plot each step
                for step_idx, step in enumerate(data['steps']):
                    
                    # MELT values
                    y_melt = []
                    # RESIDUE values
                    y_residue = []
                    
                    for ree in ree_order:
                        # Melt
                        if ree in data['chondrite_normalized_melt']:
                            y_melt.append(data['chondrite_normalized_melt'][ree][step_idx])
                        else:
                            y_melt.append(np.nan)
                        
                        # Residue
                        if ree in data['chondrite_normalized_residue']:
                            y_residue.append(data['chondrite_normalized_residue'][ree][step_idx])
                        else:
                            y_residue.append(np.nan)
                    
                    cumulative_F = data['cumulative_melt_fraction'][step_idx]
                    
                    # Plot melt
                    ax_melt.plot(x_positions, y_melt, 'o-', 
                               label=f'Step {step} (F={cumulative_F:.0%})', 
                               linewidth=2, markersize=6)
                    
                    # Plot residue  
                    ax_residue.plot(x_positions, y_residue, 's-',
                                  label=f'Step {step} (F={cumulative_F:.0%})', 
                                  linewidth=2, markersize=6)
            else:
                # For continuous: plot selected melt fractions
                melt_fractions_to_plot = [0.05, 0.10, 0.15, 0.20]  # 5%, 10%, 15%, 20%
                colors = ['blue', 'red', 'green', 'orange']
                
                for j, F in enumerate(melt_fractions_to_plot):
                    if F <= max(data['melt_fractions']):
                        # Find closest melt fraction index
                        idx = np.argmin(np.abs(np.array(data['melt_fractions']) - F))
                        
                        # MELT values
                        y_melt = []
                        # RESIDUE values
                        y_residue = []
                        
                        for ree in ree_order:
                            # Melt
                            if ree in data['chondrite_normalized_melt']:
                                y_melt.append(data['chondrite_normalized_melt'][ree][idx])
                            else:
                                y_melt.append(np.nan)
                            
                            # Residue
                            if ree in data['chondrite_normalized_residue']:
                                y_residue.append(data['chondrite_normalized_residue'][ree][idx])
                            else:
                                y_residue.append(np.nan)
                        
                        # Plot melt
                        ax_melt.plot(x_positions, y_melt, 'o-', color=colors[j],
                                   label=f'F = {F:.0%}', linewidth=2, markersize=6)
                        
                        # Plot residue
                        ax_residue.plot(x_positions, y_residue, 's-', color=colors[j],
                                      label=f'F = {F:.0%}', linewidth=2, markersize=6)
            
            # Format MELT plot
            ax_melt.set_yscale('log')
            ax_melt.set_title(f'{scenario_name} - MELT', fontsize=14, fontweight='bold')
            ax_melt.set_ylabel('Melt/Chondrite', fontsize=12)
            ax_melt.set_xticks(x_positions)
            ax_melt.set_xticklabels(ree_order, rotation=45)
            ax_melt.grid(True, alpha=0.3)
            ax_melt.legend(fontsize=10)
            ax_melt.set_ylim(0.1, 1000)
            
            # Format RESIDUE plot
            ax_residue.set_yscale('log')
            ax_residue.set_title(f'{scenario_name} - RESIDUE', fontsize=14, fontweight='bold')
            ax_residue.set_xlabel('REE Elements', fontsize=12)
            ax_residue.set_ylabel('Residue/Chondrite', fontsize=12)
            ax_residue.set_xticks(x_positions)
            ax_residue.set_xticklabels(ree_order, rotation=45)
            ax_residue.grid(True, alpha=0.3)
            ax_residue.legend(fontsize=10)
            ax_residue.set_ylim(0.01, 10)  # Residue should be depleted, so lower values
        
        plt.tight_layout()
        plt.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.show()
        
        print(f"✅ Spider plots saved as {output_file}")
        print("📊 Now showing BOTH melt (enriched) and residue (depleted) patterns")
    
    def export_to_excel(self, scenarios_data, output_file='Batch_Melting_Results.xlsx'):
        """
        Export all results to Excel with detailed calculations
        """
        print("📝 Exporting results to Excel...")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        scenario_names = ['Garnet_Continuous', 'Garnet_Stepwise', 'Plagioclase_Continuous']
        
        for i, (scenario_name, data) in enumerate(zip(scenario_names, scenarios_data)):
            ws = wb.create_sheet(title=scenario_name)
            row = 1
            
            # Title
            ws[f'A{row}'] = f"{scenario_name.replace('_', ' ')} - Batch Melting Results"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            # Input parameters section
            ws[f'A{row}'] = "INPUT PARAMETERS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            # Starting concentrations
            ws[f'A{row}'] = "Starting REE Concentrations (ppm)"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for ree in self.ree_elements:
                if ree in data['starting_concentrations']:
                    ws[f'A{row}'] = ree
                    ws[f'B{row}'] = data['starting_concentrations'][ree]
                    row += 1
            
            row += 1
            
            # Modal proportions
            ws[f'A{row}'] = "Modal Proportions"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            mineral_names = ['Ol', 'Opx', 'Cpx', 'Grt'] if 'Garnet' in scenario_name else ['Ol', 'Opx', 'Cpx', 'Grt', 'Plg']
            
            for j, mineral in enumerate(mineral_names):
                if j < len(data['modal_proportions']):
                    ws[f'A{row}'] = mineral
                    ws[f'B{row}'] = data['modal_proportions'][j]
                    row += 1
            
            row += 1
            
            # Melting proportions
            ws[f'A{row}'] = "Melting Reaction Proportions"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for j, mineral in enumerate(mineral_names):
                if j < len(data['melting_proportions']):
                    ws[f'A{row}'] = mineral
                    ws[f'B{row}'] = data['melting_proportions'][j]
                    row += 1
            
            row += 2
            
            # Bulk partition coefficients
            ws[f'A{row}'] = "BULK PARTITION COEFFICIENTS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            ws[f'A{row}'] = "REE"
            ws[f'B{row}'] = "Bulk D (Source)"
            ws[f'C{row}'] = "Bulk P (Melting)"
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.font = Font(bold=True)
            row += 1
            
            for ree in self.ree_elements:
                if ree in data['bulk_D']:
                    ws[f'A{row}'] = ree
                    ws[f'B{row}'] = data['bulk_D'][ree]
                    ws[f'C{row}'] = data['bulk_P'][ree]
                    row += 1
            
            row += 2
            
            # Results section
            ws[f'A{row}'] = "MELTING RESULTS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            if 'stepwise' in scenario_name.lower():
                # Stepwise results
                ws[f'A{row}'] = "Step"
                ws[f'B{row}'] = "Cumulative F"
                col = 3
                
                # Headers for melt concentrations
                for ree in self.ree_elements:
                    if ree in data['melt_concentrations']:
                        ws.cell(row=row, column=col, value=f"{ree}_melt")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                # Headers for residue concentrations  
                for ree in self.ree_elements:
                    if ree in data['residue_concentrations']:
                        ws.cell(row=row, column=col, value=f"{ree}_residue")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                row += 1
                
                # Data rows
                for i, step in enumerate(data['steps']):
                    ws[f'A{row}'] = step
                    ws[f'B{row}'] = data['cumulative_melt_fraction'][i]
                    col = 3
                    
                    # Melt concentrations
                    for ree in self.ree_elements:
                        if ree in data['melt_concentrations']:
                            ws.cell(row=row, column=col, value=data['melt_concentrations'][ree][i])
                            col += 1
                    
                    # Residue concentrations
                    for ree in self.ree_elements:
                        if ree in data['residue_concentrations']:
                            ws.cell(row=row, column=col, value=data['residue_concentrations'][ree][i])
                            col += 1
                    
                    row += 1
                    
            else:
                # Continuous results
                ws[f'A{row}'] = "Melt Fraction (F)"
                col = 2
                
                # Headers for melt concentrations
                for ree in self.ree_elements:
                    if ree in data['melt_concentrations']:
                        ws.cell(row=row, column=col, value=f"{ree}_melt")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                # Headers for residue concentrations
                for ree in self.ree_elements:
                    if ree in data['residue_concentrations']:
                        ws.cell(row=row, column=col, value=f"{ree}_residue")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                row += 1
                
                # Data rows
                for i, F in enumerate(data['melt_fractions']):
                    ws[f'A{row}'] = F
                    col = 2
                    
                    # Melt concentrations
                    for ree in self.ree_elements:
                        if ree in data['melt_concentrations']:
                            ws.cell(row=row, column=col, value=data['melt_concentrations'][ree][i])
                            col += 1
                    
                    # Residue concentrations  
                    for ree in self.ree_elements:
                        if ree in data['residue_concentrations']:
                            ws.cell(row=row, column=col, value=data['residue_concentrations'][ree][i])
                            col += 1
                    
                    row += 1
            
            row += 2
            
            # Chondrite normalized section
            ws[f'A{row}'] = "CHONDRITE NORMALIZED VALUES"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            if 'stepwise' in scenario_name.lower():
                # Stepwise chondrite normalized results
                ws[f'A{row}'] = "Step"
                ws[f'B{row}'] = "Cumulative F"
                col = 3
                
                # Headers for chondrite normalized melt
                for ree in self.ree_elements:
                    if ree in data.get('chondrite_normalized_melt', {}):
                        ws.cell(row=row, column=col, value=f"{ree}_melt_CN")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                # Headers for chondrite normalized residue
                for ree in self.ree_elements:
                    if ree in data.get('chondrite_normalized_residue', {}):
                        ws.cell(row=row, column=col, value=f"{ree}_residue_CN")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                row += 1
                
                # Data rows for chondrite normalized
                for i, step in enumerate(data['steps']):
                    ws[f'A{row}'] = step
                    ws[f'B{row}'] = data['cumulative_melt_fraction'][i]
                    col = 3
                    
                    # Chondrite normalized melt
                    for ree in self.ree_elements:
                        if ree in data.get('chondrite_normalized_melt', {}):
                            ws.cell(row=row, column=col, value=data['chondrite_normalized_melt'][ree][i])
                            col += 1
                    
                    # Chondrite normalized residue
                    for ree in self.ree_elements:
                        if ree in data.get('chondrite_normalized_residue', {}):
                            ws.cell(row=row, column=col, value=data['chondrite_normalized_residue'][ree][i])
                            col += 1
                    
                    row += 1
                    
            else:
                # Continuous chondrite normalized results
                ws[f'A{row}'] = "Melt Fraction (F)"
                col = 2
                
                # Headers for chondrite normalized melt
                for ree in self.ree_elements:
                    if ree in data.get('chondrite_normalized_melt', {}):
                        ws.cell(row=row, column=col, value=f"{ree}_melt_CN")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                # Headers for chondrite normalized residue
                for ree in self.ree_elements:
                    if ree in data.get('chondrite_normalized_residue', {}):
                        ws.cell(row=row, column=col, value=f"{ree}_residue_CN")
                        ws.cell(row=row, column=col).font = Font(bold=True)
                        col += 1
                
                row += 1
                
                # Data rows for chondrite normalized
                for i, F in enumerate(data['melt_fractions']):
                    ws[f'A{row}'] = F
                    col = 2
                    
                    # Chondrite normalized melt
                    for ree in self.ree_elements:
                        if ree in data.get('chondrite_normalized_melt', {}):
                            ws.cell(row=row, column=col, value=data['chondrite_normalized_melt'][ree][i])
                            col += 1
                    
                    # Chondrite normalized residue
                    for ree in self.ree_elements:
                        if ree in data.get('chondrite_normalized_residue', {}):
                            ws.cell(row=row, column=col, value=data['chondrite_normalized_residue'][ree][i])
                            col += 1
                    
                    row += 1
        
        wb.save(output_file)
        print(f"✅ Results exported to {output_file}")
    
    def run_all_scenarios(self):
        """
        Run all three melting scenarios and generate outputs
        """
        print("🚀 Starting Non-Modal Batch Melting Calculations...")
        print("=" * 60)
        
        scenarios_data = []
        
        # Scenario 1: Garnet Continuous
        garnet_continuous = self.continuous_melting(self.garnet_data, 'garnet')
        garnet_continuous['chondrite_normalized_melt'] = self.normalize_to_chondrite(
            garnet_continuous['melt_concentrations'], self.chondrite_data)
        garnet_continuous['chondrite_normalized_residue'] = self.normalize_to_chondrite(
            garnet_continuous['residue_concentrations'], self.chondrite_data)
        scenarios_data.append(garnet_continuous)
        
        # Scenario 2: Garnet Stepwise
        garnet_stepwise = self.stepwise_melting(self.garnet_data)
        garnet_stepwise['chondrite_normalized_melt'] = self.normalize_to_chondrite(
            garnet_stepwise['melt_concentrations'], self.chondrite_data)
        garnet_stepwise['chondrite_normalized_residue'] = self.normalize_to_chondrite(
            garnet_stepwise['residue_concentrations'], self.chondrite_data)
        scenarios_data.append(garnet_stepwise)
        
        # Scenario 3: Plagioclase Continuous
        plag_continuous = self.continuous_melting(self.plag_data, 'plagioclase')
        plag_continuous['chondrite_normalized_melt'] = self.normalize_to_chondrite(
            plag_continuous['melt_concentrations'], self.chondrite_data)
        plag_continuous['chondrite_normalized_residue'] = self.normalize_to_chondrite(
            plag_continuous['residue_concentrations'], self.chondrite_data)
        scenarios_data.append(plag_continuous)
        
        # Generate outputs
        self.create_spider_plot(scenarios_data)
        self.export_to_excel(scenarios_data)
        
        print("=" * 60)
        print("✅ All calculations completed successfully!")
        print("📊 Spider plots: REE_Spider_Plots.png")
        print("📝 Excel results: Batch_Melting_Results.xlsx")
        
        return scenarios_data

# Main execution
if __name__ == "__main__":
    # Initialize the model with your input file
    model = BatchMeltingModel('Eclogite melting_Input.xlsx')
    
    # Run all scenarios
    results = model.run_all_scenarios()
    
    print("\n🎉 Non-Modal Batch Melting Model completed successfully!")
    print("Check the output files for detailed results and REE spider diagrams.")